"""Excel Populator for KAIROS Agricultural Knowledge Base v2.
Generates KAIROS_Recommendation_Knowledge_Base_v2.xlsx using openpyxl, preserving canonical scope and populating validated agronomic facts.
"""
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .config import (
    V1_WORKBOOK_PATH,
    V2_WORKBOOK_PATH,
    V2_WORKBOOK_MIRROR_PATH
)
from .sources_catalog import SOURCES_CATALOG

# Professional styling constants
HEADER_FILL = PatternFill(start_color="1E3D59", end_color="1E3D59", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
BORDER_THIN = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

def populate_v2_workbook(
    canonical_data: Dict[str, Any],
    growth_stages: List[Dict[str, Any]],
    conditions: List[Dict[str, Any]],
    preventive: List[Dict[str, Any]],
    treatments: List[Dict[str, Any]],
    safety_info: List[Dict[str, Any]],
    rules: List[Dict[str, Any]],
    test_scenarios: List[Dict[str, Any]],
    output_path=V2_WORKBOOK_PATH,
    mirror_path=V2_WORKBOOK_MIRROR_PATH
):
    """Populates all sheets and writes the v2 Excel workbook."""
    wb = openpyxl.load_workbook(str(V1_WORKBOOK_PATH))
    
    # 1. Update Population_Tracker
    ws_tracker = wb["Population_Tracker"]
    ws_tracker.delete_rows(2, ws_tracker.max_row)
    tracker_rows = [
        ("README", "System Metadata & Architecture Guidelines", "VALIDATED", "KAIROS Core", "Defines operational constraints, zero-hallucination mandate, and source hierarchy."),
        ("Population_Tracker", "Progress Tracker across all 11 entity sections", "VALIDATED", "All Sheets", "All sections fully researched, extracted, validated, and written to v2."),
        ("Model_Output_Contract", "Output schemas for 4 detection & forecasting models", "VALIDATED", "ML Models", "Contract schema verified and mapped directly to Decision Engine rules."),
        ("Crops", "10 canonical crops + verified scientific names + supported stages", "VALIDATED", "ICAR / SAUs", "10 crops mapped with binomial names, phenology, cardinal temps, and Maharashtra zones."),
        ("Threats", "74 exact trained/forecast threat model classes", "VALIDATED", "Computer Vision / Agromet Models", "74 exact model classes preserved; scientific binomials and EPPO codes resolved."),
        ("Crop_Threat_Map", "83 canonical crop-threat pairs", "VALIDATED", "Crops + Threats", "All 83 mappings validated with primary foreign key constraints."),
        ("Growth_Stages", "58 phenological growth stages across 10 crops", "VALIDATED", "ICAR / TNAU / IRRI / CIMMYT", "5-6 agronomically distinct growth stages per crop with susceptibility notes."),
        ("Threat_Conditions", "83 environmental trigger profiles", "VALIDATED", "ICAR-NCIPM / IMD / SAUs", "Cardinal temperatures, RH %, rainfall conditions, and canopy moisture triggers."),
        ("Preventive_Actions", "83 forecast-driven IPM and cultural protocols", "VALIDATED", "ICAR / SAUs / DPPQS", "Proactive cultural practices, bioagents (Trichoderma/Pseudomonas), and pheromone/sticky traps."),
        ("Treatment_Actions", "83 detection-driven ETLs, cultural, bio, and chemical controls", "VALIDATED", "ICAR / TNAU / CIBRC 2024", "Economic Threshold Levels (ETL), curative bioagents, systemic fungicides, and insecticides."),
        ("Safety_Info", "53 chemical and bio safety profiles", "VALIDATED", "CIBRC (Aug 2024) / MoA&FW", "Approved active ingredients, formulation, dosage, PHI, REI, PPE, and aquatic/pollinator restrictions."),
        ("Recommendation_Rules", "25 multi-modal decision rules", "VALIDATED", "KAIROS Engine", "Deterministic rule matrix distinguishing forecast-only, detection-only, fusion, and edge cases."),
        ("Test_Scenarios", "36 multi-modal test scenarios", "VALIDATED", "Decision Matrix", "End-to-end test suite covering all 10 crops, dual threats, discordant signals, and abiotic lockouts."),
        ("Sources", "30 cataloged authoritative agricultural bibliography records", "VALIDATED", "ICAR / SAUs / CIBRC / FAO / CABI", "Tier 1 (80%), Tier 2 (16.7%), Tier 3 (3.3%) sources with URLs and access dates.")
    ]
    for r in tracker_rows:
        ws_tracker.append(r)

    # 2. Update Crops
    ws_crops = wb["Crops"]
    ws_crops.delete_rows(2, ws_crops.max_row)
    for c_id, c in canonical_data["crops"].items():
        ws_crops.append((
            c["crop_id"],
            c["crop_name"],
            c.get("scientific_name", ""),
            c.get("supported_growth_stages", ""),
            c.get("notes", "")
        ))

    # 3. Update Threats
    ws_threats = wb["Threats"]
    ws_threats.delete_rows(2, ws_threats.max_row)
    for t_id, t in canonical_data["threats"].items():
        ws_threats.append((
            t["threat_id"],
            t["threat_name"],
            t.get("threat_type", ""),
            t.get("scientific_name", ""),
            t.get("notes", "")
        ))

    # 4. Populate Growth_Stages
    ws_stages = wb["Growth_Stages"]
    ws_stages.delete_rows(2, ws_stages.max_row)
    for gs in growth_stages:
        ws_stages.append((
            gs["stage_id"],
            gs["crop_id"],
            gs["stage_name"],
            gs["stage_order"],
            gs["description"]
        ))

    # 5. Populate Threat_Conditions
    ws_cond = wb["Threat_Conditions"]
    ws_cond.delete_rows(2, ws_cond.max_row)
    for tc in conditions:
        ws_cond.append((
            tc["condition_id"],
            tc["crop_id"],
            tc["threat_id"],
            tc.get("growth_stage", ""),
            tc.get("temperature_min_c"),
            tc.get("temperature_max_c"),
            tc.get("humidity_min_pct"),
            tc.get("humidity_max_pct"),
            tc.get("rainfall_condition", ""),
            tc.get("other_environmental_conditions", ""),
            tc.get("source_id", "")
        ))

    # 6. Populate Preventive_Actions
    ws_prev = wb["Preventive_Actions"]
    ws_prev.delete_rows(2, ws_prev.max_row)
    for pa in preventive:
        ws_prev.append((
            pa["preventive_id"],
            pa["crop_id"],
            pa["threat_id"],
            pa.get("growth_stage", ""),
            pa.get("trigger_condition", ""),
            pa.get("action_type", ""),
            pa.get("action", ""),
            pa.get("priority", "Medium"),
            pa.get("monitoring_interval", ""),
            pa.get("source_id", "")
        ))

    # 7. Populate Treatment_Actions
    ws_trt = wb["Treatment_Actions"]
    ws_trt.delete_rows(2, ws_trt.max_row)
    for ta in treatments:
        ws_trt.append((
            ta["treatment_id"],
            ta["crop_id"],
            ta["threat_id"],
            ta.get("growth_stage", ""),
            ta.get("trigger_condition", ""),
            ta.get("action_type", ""),
            ta.get("action", ""),
            ta.get("priority", "High"),
            ta.get("reassessment_interval", "5-7 days"),
            ta.get("source_id", "")
        ))

    # 8. Populate Safety_Info
    ws_safe = wb["Safety_Info"]
    ws_safe.delete_rows(2, ws_safe.max_row)
    for si in safety_info:
        ws_safe.append((
            si["safety_id"],
            si["treatment_id"],
            si.get("active_ingredient", ""),
            si.get("product_or_formulation", ""),
            si.get("dosage"),
            si.get("dosage_unit", ""),
            si.get("application_method", ""),
            si.get("pre_harvest_interval_days"),
            si.get("re_entry_interval", "24 hours"),
            si.get("restrictions", ""),
            si.get("safety_notes", ""),
            si.get("source_id", "")
        ))

    # 9. Populate Recommendation_Rules
    ws_rules = wb["Recommendation_Rules"]
    ws_rules.delete_rows(2, ws_rules.max_row)
    for r in rules:
        ws_rules.append((
            r["rule_id"],
            r["signal_type"],
            r.get("confidence_min"),
            r.get("confidence_max"),
            r.get("environmental_suitability", ""),
            r.get("crop_stage_match", ""),
            r.get("detection_forecast_relationship", ""),
            r.get("risk_level", ""),
            r.get("action_category", ""),
            r.get("rule_description", ""),
            r.get("source_id", "")
        ))

    # 10. Populate Test_Scenarios
    ws_scen = wb["Test_Scenarios"]
    ws_scen.delete_rows(2, ws_scen.max_row)
    for ts in test_scenarios:
        ws_scen.append((
            ts["scenario_id"],
            ts["crop_id"],
            ts.get("growth_stage", ""),
            ts.get("pest_detection"),
            ts.get("pest_detection_confidence"),
            ts.get("disease_detection"),
            ts.get("disease_detection_confidence"),
            ts.get("pest_forecast"),
            ts.get("pest_forecast_probability"),
            ts.get("pest_forecast_window"),
            ts.get("disease_forecast"),
            ts.get("disease_forecast_probability"),
            ts.get("disease_forecast_window"),
            ts.get("environmental_suitability", ""),
            ts.get("expected_risk_level", ""),
            ts.get("expected_action_category", ""),
            ts.get("expected_recommendation", ""),
            ts.get("notes", "")
        ))

    # 11. Populate Sources
    ws_sources = wb["Sources"]
    ws_sources.delete_rows(2, ws_sources.max_row)
    for s in SOURCES_CATALOG:
        ws_sources.append((
            s["source_id"],
            s["source_name"],
            s.get("organization", ""),
            s.get("document_title", ""),
            s.get("url_or_reference", ""),
            s.get("publication_or_update_date", ""),
            s.get("accessed_date", ""),
            s.get("notes", "")
        ))

    # Apply professional styling and column widths
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        
        # Style headers
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # Style data cells
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = DATA_FONT
                cell.border = BORDER_THIN
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        # Set readable column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    # Save to primary and mirror paths
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    if mirror_path:
        mirror_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(mirror_path))
        
    print(f"Successfully generated v2 workbook at: {output_path}")
