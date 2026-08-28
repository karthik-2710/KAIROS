"""Complete Agronomic Knowledge Dataset for KAIROS KB v2.
Contains all 83 verified entity records across 10 crops, 74 threats, 58 growth stages, 83 threat conditions, 83 preventive actions, 83 treatment actions, 53 safety profiles, 25 recommendation rules, and 36 test scenarios.
"""
from typing import Dict, List, Any
import openpyxl
from .config import V2_WORKBOOK_PATH

def get_complete_agronomic_dataset():
    """Extracts the verified complete agronomic dataset directly from the authoritative repository."""
    wb = openpyxl.load_workbook(str(V2_WORKBOOK_PATH), data_only=True)
    
    # 1. Crops
    crops = {}
    for r in list(wb["Crops"].iter_rows(values_only=True))[1:]:
        if r[0]:
            crops[r[0]] = {
                "crop_id": r[0],
                "crop_name": r[1],
                "scientific_name": r[2],
                "supported_growth_stages": r[3],
                "notes": r[4]
            }

    # 2. Threats
    threats = {}
    for r in list(wb["Threats"].iter_rows(values_only=True))[1:]:
        if r[0]:
            threats[r[0]] = {
                "threat_id": r[0],
                "threat_name": r[1],
                "threat_type": r[2],
                "scientific_name": r[3],
                "notes": r[4]
            }

    # 3. Growth Stages
    growth_stages = []
    for r in list(wb["Growth_Stages"].iter_rows(values_only=True))[1:]:
        if r[0]:
            growth_stages.append({
                "stage_id": r[0],
                "crop_id": r[1],
                "stage_name": r[2],
                "stage_order": r[3],
                "description": r[4]
            })

    # 4. Threat Conditions
    conditions = []
    for r in list(wb["Threat_Conditions"].iter_rows(values_only=True))[1:]:
        if r[0]:
            conditions.append({
                "condition_id": r[0],
                "crop_id": r[1],
                "threat_id": r[2],
                "growth_stage": r[3],
                "temperature_min_c": r[4],
                "temperature_max_c": r[5],
                "humidity_min_pct": r[6],
                "humidity_max_pct": r[7],
                "rainfall_condition": r[8],
                "other_environmental_conditions": r[9],
                "source_id": r[10]
            })

    # 5. Preventive Actions
    preventive = []
    for r in list(wb["Preventive_Actions"].iter_rows(values_only=True))[1:]:
        if r[0]:
            preventive.append({
                "preventive_id": r[0],
                "crop_id": r[1],
                "threat_id": r[2],
                "growth_stage": r[3],
                "trigger_condition": r[4],
                "action_type": r[5],
                "action": r[6],
                "priority": r[7],
                "monitoring_interval": r[8],
                "source_id": r[9]
            })

    # 6. Treatment Actions
    treatments = []
    for r in list(wb["Treatment_Actions"].iter_rows(values_only=True))[1:]:
        if r[0]:
            treatments.append({
                "treatment_id": r[0],
                "crop_id": r[1],
                "threat_id": r[2],
                "growth_stage": r[3],
                "trigger_condition": r[4],
                "action_type": r[5],
                "action": r[6],
                "priority": r[7],
                "reassessment_interval": r[8],
                "source_id": r[9]
            })

    # 7. Safety Info
    safety_info = []
    for r in list(wb["Safety_Info"].iter_rows(values_only=True))[1:]:
        if r[0]:
            safety_info.append({
                "safety_id": r[0],
                "treatment_id": r[1],
                "active_ingredient": r[2],
                "product_or_formulation": r[3],
                "dosage": r[4],
                "dosage_unit": r[5],
                "application_method": r[6],
                "pre_harvest_interval_days": r[7],
                "re_entry_interval": r[8],
                "restrictions": r[9],
                "safety_notes": r[10],
                "source_id": r[11]
            })

    # 8. Recommendation Rules
    rules = []
    for r in list(wb["Recommendation_Rules"].iter_rows(values_only=True))[1:]:
        if r[0]:
            rules.append({
                "rule_id": r[0],
                "signal_type": r[1],
                "confidence_min": r[2],
                "confidence_max": r[3],
                "environmental_suitability": r[4],
                "crop_stage_match": r[5],
                "detection_forecast_relationship": r[6],
                "risk_level": r[7],
                "action_category": r[8],
                "rule_description": r[9],
                "source_id": r[10]
            })

    # 9. Test Scenarios
    test_scenarios = []
    for r in list(wb["Test_Scenarios"].iter_rows(values_only=True))[1:]:
        if r[0]:
            test_scenarios.append({
                "scenario_id": r[0],
                "crop_id": r[1],
                "growth_stage": r[2],
                "pest_detection": r[3],
                "pest_detection_confidence": r[4],
                "disease_detection": r[5],
                "disease_detection_confidence": r[6],
                "pest_forecast": r[7],
                "pest_forecast_probability": r[8],
                "pest_forecast_window": r[9],
                "disease_forecast": r[10],
                "disease_forecast_probability": r[11],
                "disease_forecast_window": r[12],
                "environmental_suitability": r[13],
                "expected_risk_level": r[14],
                "expected_action_category": r[15],
                "expected_recommendation": r[16],
                "notes": r[17]
            })

    return {
        "canonical_data": {"crops": crops, "threats": threats},
        "growth_stages": growth_stages,
        "conditions": conditions,
        "preventive": preventive,
        "treatments": treatments,
        "safety_info": safety_info,
        "rules": rules,
        "test_scenarios": test_scenarios
    }
