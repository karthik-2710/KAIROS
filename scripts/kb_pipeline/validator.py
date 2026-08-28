"""Deterministic Python Validation Engine for KAIROS Knowledge Base.
Enforces strict schema constraints, foreign key referential integrity, numerical ranges, and chemical safety checks.
"""
from typing import Dict, List, Any, Tuple
import openpyxl

class KBValidator:
    def __init__(self, workbook_path: str):
        self.workbook_path = str(workbook_path)
        self.wb = openpyxl.load_workbook(self.workbook_path, data_only=True)
        self.errors = []
        self.warnings = []

    def validate_all(self) -> Tuple[bool, List[str], List[str]]:
        """Runs the complete suite of integrity checks."""
        self.errors = []
        self.warnings = []

        crops = self._validate_crops()
        threats = self._validate_threats()
        sources = self._validate_sources()
        mappings = self._validate_mappings(crops, threats)
        treatments = self._validate_treatments(crops, threats, sources)
        self._validate_growth_stages(crops)
        self._validate_conditions(crops, threats, sources)
        self._validate_preventive(crops, threats, sources)
        self._validate_safety(treatments, sources)
        self._validate_rules(sources)
        self._validate_test_scenarios(crops)

        is_valid = len(self.errors) == 0
        return is_valid, self.errors, self.warnings

    def _validate_crops(self) -> Dict[str, str]:
        crops = {}
        ws = self.wb["Crops"]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            crop_id = str(row[0]).strip()
            crop_name = str(row[1]).strip() if row[1] else ""
            if crop_id in crops:
                self.errors.append(f"Duplicate crop_id: {crop_id}")
            crops[crop_id] = crop_name
        if len(crops) != 10:
            self.errors.append(f"Expected 10 canonical crops, found {len(crops)}")
        return crops

    def _validate_threats(self) -> Dict[str, str]:
        threats = {}
        ws = self.wb["Threats"]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            threat_id = str(row[0]).strip()
            threat_name = str(row[1]).strip() if row[1] else ""
            if threat_id in threats:
                self.errors.append(f"Duplicate threat_id: {threat_id}")
            threats[threat_id] = threat_name
        if len(threats) != 74:
            self.errors.append(f"Expected 74 canonical threats, found {len(threats)}")
        return threats

    def _validate_sources(self) -> Dict[str, str]:
        sources = {}
        ws = self.wb["Sources"]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            source_id = str(row[0]).strip()
            source_name = str(row[1]).strip() if row[1] else ""
            if source_id in sources:
                self.errors.append(f"Duplicate source_id: {source_id}")
            sources[source_id] = source_name
        if len(sources) < 20:
            self.warnings.append(f"Low source catalog count: {len(sources)}")
        return sources

    def _validate_mappings(self, crops: Dict[str, str], threats: Dict[str, str]) -> List[Tuple[str, str]]:
        mappings = []
        ws = self.wb["Crop_Threat_Map"]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            map_id = str(row[0]).strip()
            crop_id = str(row[1]).strip() if row[1] else ""
            threat_id = str(row[2]).strip() if row[2] else ""
            
            if crop_id not in crops:
                self.errors.append(f"Crop_Threat_Map {map_id} references invalid crop_id: {crop_id}")
            if threat_id not in threats:
                self.errors.append(f"Crop_Threat_Map {map_id} references invalid threat_id: {threat_id}")
            mappings.append((crop_id, threat_id))
            
        if len(mappings) != 83:
            self.errors.append(f"Expected 83 canonical crop-threat mappings, found {len(mappings)}")
        return mappings

    def _validate_growth_stages(self, crops: Dict[str, str]):
        ws = self.wb["Growth_Stages"]
        stage_ids = set()
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            stage_id = str(row[0]).strip()
            crop_id = str(row[1]).strip() if row[1] else ""
            stage_name = str(row[2]).strip() if row[2] else ""
            stage_order = row[3]

            if stage_id in stage_ids:
                self.errors.append(f"Duplicate stage_id: {stage_id}")
            stage_ids.add(stage_id)

            if crop_id not in crops:
                self.errors.append(f"Growth_Stages {stage_id} references unknown crop_id: {crop_id}")
            if not isinstance(stage_order, (int, float)) or stage_order < 1:
                self.errors.append(f"Growth_Stages {stage_id} invalid stage_order: {stage_order}")

    def _validate_conditions(self, crops: Dict[str, str], threats: Dict[str, str], sources: Dict[str, str]):
        ws = self.wb["Threat_Conditions"]
        cond_ids = set()
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            cond_id = str(row[0]).strip()
            crop_id = str(row[1]).strip() if row[1] else ""
            threat_id = str(row[2]).strip() if row[2] else ""
            t_min = row[4]
            t_max = row[5]
            rh_min = row[6]
            rh_max = row[7]
            source_id = str(row[10]).strip() if row[10] else ""

            if cond_id in cond_ids:
                self.errors.append(f"Duplicate condition_id: {cond_id}")
            cond_ids.add(cond_id)

            if crop_id not in crops:
                self.errors.append(f"Threat_Conditions {cond_id} invalid crop_id: {crop_id}")
            if threat_id not in threats:
                self.errors.append(f"Threat_Conditions {cond_id} invalid threat_id: {threat_id}")
            if not source_id or source_id not in sources:
                self.errors.append(f"Threat_Conditions {cond_id} missing/invalid source_id: {source_id}")

            # Numerical consistency
            if t_min is not None and t_max is not None:
                if isinstance(t_min, (int, float)) and isinstance(t_max, (int, float)):
                    if t_min > t_max:
                        self.errors.append(f"Threat_Conditions {cond_id} t_min ({t_min}) > t_max ({t_max})")
            if rh_min is not None:
                if isinstance(rh_min, (int, float)) and (rh_min < 0 or rh_min > 100):
                    self.errors.append(f"Threat_Conditions {cond_id} invalid rh_min: {rh_min}")
            if rh_max is not None:
                if isinstance(rh_max, (int, float)) and (rh_max < 0 or rh_max > 100):
                    self.errors.append(f"Threat_Conditions {cond_id} invalid rh_max: {rh_max}")

    def _validate_preventive(self, crops: Dict[str, str], threats: Dict[str, str], sources: Dict[str, str]):
        ws = self.wb["Preventive_Actions"]
        prev_ids = set()
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            prev_id = str(row[0]).strip()
            crop_id = str(row[1]).strip() if row[1] else ""
            threat_id = str(row[2]).strip() if row[2] else ""
            action = str(row[6]).strip() if row[6] else ""
            source_id = str(row[9]).strip() if row[9] else ""

            if prev_id in prev_ids:
                self.errors.append(f"Duplicate preventive_id: {prev_id}")
            prev_ids.add(prev_id)

            if crop_id not in crops:
                self.errors.append(f"Preventive_Actions {prev_id} invalid crop_id: {crop_id}")
            if threat_id not in threats:
                self.errors.append(f"Preventive_Actions {prev_id} invalid threat_id: {threat_id}")
            if not action:
                self.errors.append(f"Preventive_Actions {prev_id} action is empty")
            if not source_id or source_id not in sources:
                self.errors.append(f"Preventive_Actions {prev_id} missing/invalid source_id: {source_id}")

    def _validate_treatments(self, crops: Dict[str, str], threats: Dict[str, str], sources: Dict[str, str]) -> Dict[str, Tuple[str, str]]:
        treatments = {}
        ws = self.wb["Treatment_Actions"]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            trt_id = str(row[0]).strip()
            crop_id = str(row[1]).strip() if row[1] else ""
            threat_id = str(row[2]).strip() if row[2] else ""
            action = str(row[6]).strip() if row[6] else ""
            source_id = str(row[9]).strip() if row[9] else ""

            if trt_id in treatments:
                self.errors.append(f"Duplicate treatment_id: {trt_id}")
            treatments[trt_id] = (crop_id, threat_id)

            if crop_id not in crops:
                self.errors.append(f"Treatment_Actions {trt_id} invalid crop_id: {crop_id}")
            if threat_id not in threats:
                self.errors.append(f"Treatment_Actions {trt_id} invalid threat_id: {threat_id}")
            if not action:
                self.errors.append(f"Treatment_Actions {trt_id} action is empty")
            if not source_id or source_id not in sources:
                self.errors.append(f"Treatment_Actions {trt_id} missing/invalid source_id: {source_id}")

            # Guardrail: Abiotic disorders must NOT prescribe chemical fungicides or insecticides
            threat_name = threats.get(threat_id, "")
            if any(ab in threat_name for ab in ["Herbicide Growth Damage", "Leaf Redding", "Leaf Variegation"]):
                action_lower = action.lower()
                # Check for positive chemical prescriptions (excluding explicit 'do not' prohibitions)
                positive_chemical_cues = ["spray propiconazole", "spray imidacloprid", "spray hexaconazole", "spray thiamethoxam", "spray chlorpyrifos", "spray fipronil", "spray emamectin"]
                for prob in positive_chemical_cues:
                    if prob in action_lower:
                        self.errors.append(f"Abiotic threat {threat_name} ({trt_id}) contains prohibited chemical recommendation: {prob}")

        return treatments

    def _validate_safety(self, treatments: Dict[str, Tuple[str, str]], sources: Dict[str, str]):
        ws = self.wb["Safety_Info"]
        safety_ids = set()
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            safety_id = str(row[0]).strip()
            trt_id = str(row[1]).strip() if row[1] else ""
            active_ing = str(row[2]).strip() if row[2] else ""
            dosage = row[4]
            phi = row[7]
            source_id = str(row[11]).strip() if row[11] else ""

            if safety_id in safety_ids:
                self.errors.append(f"Duplicate safety_id: {safety_id}")
            safety_ids.add(safety_id)

            if trt_id not in treatments:
                self.errors.append(f"Safety_Info {safety_id} invalid treatment_id: {trt_id}")
            if not source_id or source_id not in sources:
                self.errors.append(f"Safety_Info {safety_id} missing/invalid source_id: {source_id}")

            if dosage is not None and not isinstance(dosage, (int, float)):
                self.warnings.append(f"Safety_Info {safety_id} dosage is non-numeric: {dosage}")
            if phi is not None and not isinstance(phi, (int, float)):
                self.warnings.append(f"Safety_Info {safety_id} PHI is non-numeric: {phi}")

    def _validate_rules(self, sources: Dict[str, str]):
        ws = self.wb["Recommendation_Rules"]
        rule_ids = set()
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            rule_id = str(row[0]).strip()
            conf_min = row[2]
            conf_max = row[3]
            source_id = str(row[10]).strip() if row[10] else ""

            if rule_id in rule_ids:
                self.errors.append(f"Duplicate rule_id: {rule_id}")
            rule_ids.add(rule_id)

            if conf_min is not None and isinstance(conf_min, (int, float)) and (conf_min < 0 or conf_min > 1):
                self.errors.append(f"Rule {rule_id} conf_min out of range 0-1: {conf_min}")
            if conf_max is not None and isinstance(conf_max, (int, float)) and (conf_max < 0 or conf_max > 1):
                self.errors.append(f"Rule {rule_id} conf_max out of range 0-1: {conf_max}")
            if not source_id or source_id not in sources:
                self.errors.append(f"Rule {rule_id} missing/invalid source_id: {source_id}")

    def _validate_test_scenarios(self, crops: Dict[str, str]):
        ws = self.wb["Test_Scenarios"]
        scen_ids = set()
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            scen_id = str(row[0]).strip()
            crop_id = str(row[1]).strip() if row[1] else ""
            expected_risk = str(row[14]).strip() if row[14] else ""
            expected_rec = str(row[16]).strip() if row[16] else ""

            if scen_id in scen_ids:
                self.errors.append(f"Duplicate scenario_id: {scen_id}")
            scen_ids.add(scen_id)

            if crop_id not in crops:
                self.errors.append(f"Test_Scenarios {scen_id} invalid crop_id: {crop_id}")
            if not expected_risk:
                self.errors.append(f"Test_Scenarios {scen_id} missing expected_risk_level")
            if not expected_rec:
                self.errors.append(f"Test_Scenarios {scen_id} missing expected_recommendation")
