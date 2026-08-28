"""Task Builder for KAIROS Agricultural Research Pipeline.
Reads canonical scope from KAIROS_Recommendation_Knowledge_Base_v1.xlsx and generates 83 atomic research tasks.
"""
from typing import Dict, List, Any
import openpyxl
from .config import V1_WORKBOOK_PATH

def load_canonical_scope(v1_path=V1_WORKBOOK_PATH) -> Dict[str, Any]:
    """Loads Crops, Threats, and Crop_Threat_Map sheets from v1 Excel workbook."""
    wb = openpyxl.load_workbook(str(v1_path), data_only=True)
    
    # Load Crops
    crops = {}
    for row in wb["Crops"].iter_rows(values_only=True):
        if row[0] and row[0] != "crop_id":
            crops[row[0]] = {
                "crop_id": str(row[0]).strip(),
                "crop_name": str(row[1]).strip() if row[1] else "",
                "scientific_name": str(row[2]).strip() if row[2] else "",
                "supported_growth_stages": str(row[3]).strip() if row[3] else "",
                "notes": str(row[4]).strip() if row[4] else ""
            }
            
    # Load Threats
    threats = {}
    for row in wb["Threats"].iter_rows(values_only=True):
        if row[0] and row[0] != "threat_id":
            threats[row[0]] = {
                "threat_id": str(row[0]).strip(),
                "threat_name": str(row[1]).strip() if row[1] else "",
                "threat_type": str(row[2]).strip() if row[2] else "",
                "scientific_name": str(row[3]).strip() if row[3] else "",
                "notes": str(row[4]).strip() if row[4] else ""
            }
            
    # Load Mappings
    mappings = []
    for row in wb["Crop_Threat_Map"].iter_rows(values_only=True):
        if row[0] and row[0] != "map_id":
            mappings.append({
                "map_id": str(row[0]).strip(),
                "crop_id": str(row[1]).strip(),
                "threat_id": str(row[2]).strip(),
                "relevance": str(row[3]).strip() if row[3] else "",
                "notes": str(row[4]).strip() if row[4] else ""
            })
            
    return {
        "crops": crops,
        "threats": threats,
        "mappings": mappings
    }

def generate_research_tasks(canonical_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Constructs 83 atomic research tasks with targeted search queries and agronomic context."""
    crops = canonical_data["crops"]
    threats = canonical_data["threats"]
    mappings = canonical_data["mappings"]
    
    tasks = []
    for m in mappings:
        crop_id = m["crop_id"]
        threat_id = m["threat_id"]
        crop = crops.get(crop_id, {})
        threat = threats.get(threat_id, {})
        
        crop_name = crop.get("crop_name", "")
        threat_name = threat.get("threat_name", "")
        threat_type = threat.get("threat_type", "")
        
        # Build search queries
        clean_threat_name = threat_name
        if "PESGL_DREFCR0" in threat_name:
            clean_threat_name = "Bajra Exserohilum rostrata leaf blight"
        elif "PESGL_MOESBU" in threat_name:
            clean_threat_name = "Bajra smut Moesziomyces bullatus"
        elif "PESGL_PYRISP" in threat_name:
            clean_threat_name = "Bajra blast Pyricularia grisea"
        elif "PESGL_SCLPGR" in threat_name:
            clean_threat_name = "Bajra downy mildew green ear Sclerospora graminicola"
        elif "Herbicide Growth Damage" in threat_name:
            clean_threat_name = "Cotton herbicide injury 2,4-D glyphosate drift"
        elif "Leaf Redding" in threat_name:
            clean_threat_name = "Bt cotton leaf reddening physiological disorder"
        elif "Leaf Variegation" in threat_name:
            clean_threat_name = "Cotton leaf variegation somatic chimerism micronutrient chlorosis"
        
        search_queries = [
            f'"{crop_name}" "{clean_threat_name}" site:icar.gov.in',
            f'"{crop_name}" "{clean_threat_name}" site:tnau.ac.in',
            f'"{crop_name}" "{clean_threat_name}" "integrated pest management" India',
            f'"{crop_name}" "{clean_threat_name}" pesticide recommendation CIBRC India'
        ]
        
        task = {
            "map_id": m["map_id"],
            "crop_id": crop_id,
            "crop_name": crop_name,
            "threat_id": threat_id,
            "threat_name": threat_name,
            "threat_type": threat_type,
            "clean_threat_name": clean_threat_name,
            "search_queries": search_queries
        }
        tasks.append(task)
        
    return tasks
